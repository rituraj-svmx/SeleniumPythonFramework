import time

import openpyxl
from selenium.webdriver.common.by import By

from pageObject.UploadFile import UploadFile
from tests.conftest import driver
from utilities.BaseClass import BaseClass




class TestUploadFile(BaseClass):

    def test_file_upload(self):
        file_path = "C:/Users/XW334BQ/Downloads/download.xlsx"
        fruit_name = "Apple"
        column_name = "Price"
        new_price = "999"
        log = self.getlogger()
        upload_file = UploadFile(self.driver)
        upload_file.download_file().click()
        log.info("file downloaded successfully")
        time.sleep(7)
        self.driver.implicitly_wait(10)
        updated_data = self.update_data(file_path, fruit_name, column_name.lower(), new_price)
        file = upload_file.upload_file()
        file.send_keys(file_path)
        self.verify_toast_message(upload_file.toast_locator)
        log.info("file uploaded successfully")
        column = self.driver.find_element(By.XPATH, "//div[text()='" + column_name + "']/parent::div").get_attribute(
            "data-column-id")
        price = self.driver.find_element(By.XPATH,
                                         "//div[text()='" + fruit_name + "']/parent::div//following-sibling::div[@data-column-id='" + column + "']/child::div").text

        assert price == updated_data
        log.info("Actual price is "+ price + " equal to updated price "+ updated_data)

    def update_data(self, file_path, fruitname, col_name, price):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        rows = sheet.max_row
        columns = sheet.max_column
        dict = {}
        for i in range(1, columns + 1):
                if sheet.cell(row=1, column=i).value == col_name:
                    dict["col"] = i
                    break


        for i in range(1, rows + 1):
            for j in range(1, columns + 1):
                if sheet.cell(row=i, column=j).value == fruitname:
                    dict["row"] = i
                    break

        sheet.cell(row=dict['row'],column=dict['col']).value = price
        workbook.save(file_path)
        return sheet.cell(row=dict['row'],column=dict['col']).value











