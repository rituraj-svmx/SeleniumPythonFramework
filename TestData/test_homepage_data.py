import openpyxl


class TestHomePageData:
        test_data_homepage = [{"firstname": "Rituraj", "email": "Rituraj@gmail.com", "gender": "Male"}]

        @staticmethod
        def get_testdata(test_case_name):
                workbook = openpyxl.load_workbook("C:\\Users\\XW334BQ\\PycharmProjects\\usersdata.xlsx")
                sheet = workbook.active
                usersData = {}
                rows = sheet.max_row
                columns = sheet.max_column
                for row in range(1, rows + 1):
                        if sheet.cell(row=row, column=1).value == test_case_name:
                                for column in range(2, columns + 1):
                                        usersData[sheet.cell(row=1, column=column).value] = sheet.cell(row=row,
                                                                                                       column=column).value

                return [usersData]
